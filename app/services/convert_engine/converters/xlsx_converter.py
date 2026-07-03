"""XLSX converter."""

from pathlib import Path
from typing import List

from app.services.convert_engine.converters.base import BaseConverter
from app.services.convert_engine.utils import ConversionError, estimate_row_pages, estimate_text_pages, non_empty_lines


class XlsxConverter(BaseConverter):
    file_type = "xlsx"

    def supported_targets(self) -> List[str]:
        return ["pdf", "txt", "hwpx"]

    def extract_text(self, source_path: Path) -> str:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ConversionError("XLSX conversion requires openpyxl.") from exc

        workbook = load_workbook(source_path, data_only=True, read_only=True)
        try:
            sections = []
            for sheet in workbook.worksheets:
                rows = []
                for row in sheet.iter_rows(values_only=True):
                    values = ["" if value is None else str(value) for value in row]
                    if any(value.strip() for value in values):
                        rows.append("\t".join(values).rstrip())
                if rows:
                    sections.append(f"[{sheet.title}]\n" + "\n".join(rows))
            return "\n\n".join(sections).strip()
        finally:
            workbook.close()

    def write_text(self, text: str, output_path: Path, title: str = "") -> None:
        try:
            from openpyxl import Workbook
        except ImportError as exc:
            raise ConversionError("XLSX conversion requires openpyxl.") from exc

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Converted"
        sheet.append(["Line", "Text"])
        for index, line in enumerate(non_empty_lines(text), start=1):
            sheet.append([index, line])
        workbook.save(output_path)

    def page_count(self, source_path: Path, text: str) -> int:
        try:
            from openpyxl import load_workbook

            workbook = load_workbook(source_path, read_only=True, data_only=True)
            try:
                total_rows = sum(max(0, sheet.max_row or 0) for sheet in workbook.worksheets)
            finally:
                workbook.close()
            return estimate_row_pages(total_rows)
        except Exception:
            return estimate_text_pages(text)
